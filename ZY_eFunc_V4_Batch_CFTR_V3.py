# functions to import for main py
# next, put them inside subplot(1,2): done
#   09/23/2020: should be perfect to save in excel with protocol name
#   and don't repeatly write if data exit
# next: batch analysis: done on 09/25/2020
# next: need to get the decay tau of density trace
# %reset
# like clear in MatLab
# next: make it better for multiple vols
import os
import tkinter as tk
from tkinter import filedialog
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pyabf
import scipy.optimize as opt # curve_fit
from scipy.signal import find_peaks
from openpyxl import load_workbook
from operator import itemgetter


# def load_abf():
#     if st.button('Upload file'):
#         root = tk.Tk()
#         root.withdraw()
#         file_path = filedialog.askopenfilename()
#         abf = pyabf.ABF(file_path)
#         return abf

# exp fit
def func(x, a, tau, c):
    return a * np.exp(-x/tau) + c

def plotRaw(abf, target_vol, rasterized=False):
    cm = plt.get_cmap("winter")
    colors = [cm(x / abf.sweepCount) for x in abf.sweepList]
    # colors.reverse()
    rampT1 = rampP1 * abf.dataSecPerPoint  # start time in s
    rampT2 = rampP2 * abf.dataSecPerPoint  # stop time in s
    targetP = (target_vol - rampV1) * (rampP2 - rampP1) / (rampV2 - rampV1) + rampP1
    targetTimeSec = targetP * abf.dataSecPerPoint
    fig = plt.figure(figsize=(8, 8))
    fig.suptitle(abf.abfID)
    # plot traces
    ax1 = fig.add_subplot(211)
    for sweepNumber in abf.sweepList:
        abf.setSweep(sweepNumber)
        # i1, i2 = 0, int(abf.dataRate * 1)
        # dataX = abf.sweepX[i1:i2] + .025 * sweepNumber
        # dataY = abf.sweepY[i1:i2] + 15 * sweepNumber
        ax1.plot(abf.sweepX, abf.sweepY, color=colors[sweepNumber], alpha=.5, rasterized=rasterized)
    ax1.set_ylabel(abf.sweepLabelY)
    ax1.legend([f'{abf.sweepCount} sweeps'], handletextpad=0, handlelength=0)

    # plot protocol
    ax2 = fig.add_subplot(212)
    abf.setSweep(1)
    ax2.plot(abf.sweepX, abf.sweepC, color='r', alpha=.5, rasterized=rasterized)
    ax2.set_ylabel(abf.sweepLabelC)
    ax2.set_xlabel(abf.sweepLabelX)

    # plt.gca().axis('off')
    for ax in fig.get_axes():
        ax.grid(alpha=.5, ls='--')
        ax.axvspan(rampT1, rampT2, color='r', alpha=.3, lw=0)
        ax.axvline(targetTimeSec, color='blue', alpha=.3, lw=2)
        ax.set_xlim(0, max(abf.sweepX))
        # ax.label_outer()
    return fig
def getCurrentDensity(abf, vol, bl):
    # get the raw I(target_vol)
    pt = (vol - rampV1) * (rampP2 - rampP1) / (rampV2 - rampV1) + rampP1
    pt = round(pt)
    currents = []
    for sweep in abf.sweepList:
        abf.setSweep(sweep)
        # get current at the voltage of v (10 points)
        currents.append(np.average(abf.sweepY[pt - 5:pt + 5]))  # ave back to 10 points

    # let's find baseline (min) here
    if not bl or bl[-1] > abf.sweepCount+1:
        print(f'max sweep is {abf.sweepCount}, out of sweep range')
        # bl = [int(i) for i in input(f'The baseline sweepNum range under {abf.sweepCount} (,):').split(',')]
        currents_abs = list(map(abs, currents))
        bl_rindex = len(currents_abs) - currents_abs[::-1].index(min(currents_abs[170:])) - 1 #inh 172 applied at sweep170
        bl = [bl_rindex-2, bl_rindex+2] # at most 8 sec
    [bl_start, bl_end] = bl

    # baseline range
    current_BL = np.average(currents[bl_start:bl_end])
    currentDensity = [(i - current_BL) / mem_cap for i in currents]

    return currentDensity, bl


def plotDensity(data, xcol, ycol, sweepRange, ax, target_vol, showpeak=True):
    # plot x=xCol, y=yCol, from pos1 to pos2
    if not sweepRange or len(sweepRange)!=2:
        sweepRange = [int(i) for i in input(f'The sweepNum range to plot(start,end<{len(data)}):').split(',')]
    elif sweepRange[1] > len(data):
        sweepRange[1] = len(data)
    [pos1, pos2] = sweepRange
    data.iloc[pos1:pos2].plot(x=xcol, y=ycol, ax=ax, label=f'{round(target_vol)} mV')
    ax.set_xlabel(abf.sweepLabelX)
    ax.set_ylabel('Current Density (pA/pF)')
    # ax.legend([f'{round(target_vol)} mV'], loc='lower left')
    ax.grid(alpha=.5, ls='--')

    # ax.set_title(abf.abfID

    # peak
    # peakIndex = find_peaks(abs(data[ycol]), height=10, distance=30)[0]
    # peakValue = data[ycol][peakIndex].values.tolist()
    # mostly one peak with unknown height, so to max
    peakIndex = np.array([abs(data[ycol]).tolist().index(max(abs(data[ycol][30:])))]) #get first max during drug
    # peakValue = data[ycol][peakIndex - 1:peakIndex + 1].values.tolist()
    peakValue = [np.average(data[ycol][j-1:j+1]) for j in peakIndex]
    peakValue = [round(y, 2) for y in peakValue]
    print(peakIndex)
    # fit the first current from peakIndex[0]+3 to currentFitTill
    # x0 = data[xcol][peakIndex[0]+3]
    # xFit = data[xcol][peakIndex[0]+2:currentFitTill]
    # yFit = data[ycol][peakIndex[0]+2:currentFitTill]
    # init_guess = (max(yFit), 1, max(yFit)-min(yFit))
    # popt = opt.curve_fit(func, xFit-x0, yFit, p0=init_guess)[0]
    # print(f'current fit as {popt}')
    # tau = popt[1]
    #
    # # rising rate from start (16+2) to 25 sweep: 17:25 in python
    # r1, r2 = [int(k) for k in input(f'The sweepNum range to calculate slope):').split(',')]
    # riseRate = round(abs((data[ycol][r2]-data[ycol][r1])/(data[xcol][r2]-data[xcol][r1])), 2) #pA/pF/s
    if showpeak:
        # t = np.linspace(x0, max(xFit), 2000)
        # ax.plot(t, func(t-x0, *popt), '--r',
        #         label='Fit:\n$y = %0.2f e^{-x/%0.2f} + %0.2f$' % (popt[0],popt[1],popt[2]))

        # ax.plot(data[xcol][r1:r2], data[ycol][r1:r2], '--r')
        # ax.text(data[xcol][(r1+r2)//2], data[ycol][(r1+r2)//2], 'rate=' + str(riseRate) + ' pA*s/pF')
        for j, peakDen in zip(peakIndex, peakValue):
            x = data[xcol][j]
            # y = data[ycol][j]
            y = peakDen # use 2 peak to average the peak value
            # y = round(y, 2)
            ax.axhline(y, color='red', alpha=.3, lw=2)
            # ax.plot(data[xcol][j-1:j+1], data[ycol][j-1:j+1], '-')
            ax.text(x + 25, y, 's' + str(j + 1) + ',' + str(y))
    ax.legend(loc='lower right')
    return peakIndex, peakValue#, tau, riseRate # sweepList actually starts from 0

def getIV(abfFile, sweepNum, bl): # make it for single peak first, try to average 3 latter
    if not bl or bl[1] > abf.sweepCount+1:
        print(f'max sweep is {abf.sweepCount}, out of sweep range')
        bl = [int(i) for i in input(f'The baseline sweepNum range under {abf.sweepCount} (,):').split(',')]
    baselineSweeps = []
    for sweep in range(bl[0], bl[1]):
        if sweep < abf.sweepCount:
            abfFile.setSweep(sweep)
            baselineSweeps.append(abfFile.sweepY)
    baselineAverage = np.average(baselineSweeps, 0)
    sweeps = pd.DataFrame()
    iv_sweeps = []
    for i, j in enumerate(sweepNum):
        label = f'Peak {i + 1} at Sweep {j + 1}'
        colname = f'{abfFile.abfID}_S{j+1}'
        for k in range(j-2,j+1):
            try:
                abfFile.setSweep(k)
                iv_sweeps.append(abfFile.sweepY)
            except:
                continue
        sweeps[colname] = (np.average(iv_sweeps, 0) - baselineAverage)[rampP1:rampP2] / mem_cap
    return sweeps

def plotIV(data, xcol, ycol, ax, showfit=True):
    data.plot(x=xcol, y=ycol, ax=ax)
    # calculate reversal potential
    Efit = []

    for i, y in enumerate(ycol):
        # fit with np
        try:
            fx = np.poly1d(np.polyfit(data[xcol], data[y], nFit))
            print(fx)
            t = np.linspace(-100, 100, rampP2 - rampP1)
            x_cross = opt.brentq(fx, -130, 130)
            print('Efit is ')
            print(x_cross)
            Efit.append(round(x_cross, 2))  # reversal potential by polyfit
            # ax = plt.gca()
            ax.text(x_cross, i + 1, f'RP-{i}={round(x_cross, 2)} mV')
        except:
            print(f'Fail to fit the IV curve for {abf.abfID}')
            continue

        if showfit:
            ax.plot(t, fx(t), '-',
                    label='Fit:\n $y = %0.2e x^{4}+%0.2e x^{3}+%0.2e x^{3}+%0.2e x+%0.2e$'
                          %(fx[4],fx[3],fx[2],fx[1],fx[0]))

    ax.title.set_text('I-V curve')
    ax.set_xlabel(abf.sweepLabelC)
    ax.set_ylabel('Current Density (pA/pF)')
    ax.grid(alpha=.5, ls='--')

    # ax.xaxis.set_label_coords(.85, 0.82)
    ax.spines['left'].set_position(('data', 0))
    ax.spines['bottom'].set_position(('data', 0))
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend()
    # plt.show()
    return Efit
def runOneAbf(filename):
    global abf, rampP1, rampP2, rampV1, rampV2, mem_cap # can make a class open() to set them as abf.rampP1
    abf = pyabf.ABF(file_path)
    # or ask for baselineSweeps
    # baselineSweeps = [80, 89]
    # target_vol = -100 # mV
    # get parameters
    epochNumber = abf.sweepEpochs.types.index('Ramp')  # epoch for ramp
    rampP1 = round(abf.sweepEpochs.p1s[epochNumber])
    rampP2 = round(abf.sweepEpochs.p2s[epochNumber])
    rampV1 = abf.sweepEpochs.levels[epochNumber - 1]
    rampV2 = abf.sweepEpochs.levels[epochNumber]
    # rampT1 = rampP1 * abf.dataSecPerPoint  # start time in s
    # get telegraph mem_cap
    tele_cap = float(abf.headerText.split('fTelegraphMembraneCap = [')[1].split(']')[0])  # should be a better way
    if tele_cap > 5:
        mem_cap = tele_cap
    elif 0.5 < tele_cap <= 5:
        mem_cap = tele_cap * 10  # new setup with 10 times smaller teleCap
    else:
        print(f'mem_cap in header as {tele_cap} is too small, may not have tele on')
        mem_cap = float(input('Enter the mem_cap from notebook:'))
    mem_cap = round(mem_cap, 2)
    print(f'mem_cap is {mem_cap} pf')

    # create three df for output inside excel
    for target_vol in target_vols:
        pdfName = os.path.join(outPath, f'{abf.abfID}_{target_vol}.pdf')
        with PdfPages(pdfName) as pdf:
            # 20201126: don't need to save raw, for test only
            # fig0 = plotRaw(abf, target_vol, rasterized=True) # true to reduce size
            # pdf.savefig(fig0)
            # #plt.pause(5)
            # plt.close()

            df_density = pd.DataFrame({'sweepNumber': abf.sweepList,
                                   'sweepTimeSec': abf.sweepTimesSec.tolist()})
            df_iv = pd.DataFrame({'voltage': abf.sweepC[rampP1:rampP2]})


            # get currentDensity at voltage, baseline subtracted and normalized to mem_cap
            # densityName = f'{abf.abfID}_{target_vol}'
            df_density[abf.abfID], bl_adjusted = getCurrentDensity(abf, vol=target_vol, bl=baselineSweeps)
            # df_density[densityName] = getCurrentDensity(abf, vol=target_vol, bl=baselineSweeps)

            # plot density, get peakIndex information
            fig, ax = plt.subplots(2, 1, figsize=(8, 8))

            peakIndex, peakValue = plotDensity(data=df_density,
                                            xcol='sweepTimeSec',
                                            ycol=abf.abfID,
                                            sweepRange=sweepRange,
                                            ax=ax[0],
                                            target_vol=target_vol,
                                            showpeak=True)
            # print(peakIndex)
            # print(peakValue)
            # ax[0].legend([f'{round(target_vol)} mV'], loc='lower left')
            # print('peak sweep are: %s' % (peakIndex))
            print(f'peak sweep (from 0) and value for {target_vol}: %s' % ([peakIndex, peakValue]))

            # get IV for each peak
            sweepIV = getIV(abf, sweepNum=peakIndex, bl=bl_adjusted)
            df_iv = pd.concat([df_iv, sweepIV], axis=1)

            # plot IV at sweepNum, get reversal potential
            Efit = plotIV(df_iv,
                              xcol='voltage',
                              ycol=sweepIV.columns,
                              showfit=False,
                              ax=ax[1])
            # print('Reversal potential for each peak: %s' % (E0))
            # print('Reversal potential by polyfit: %s' % (Efit))
            fig.suptitle(abf.abfID)
            plt.tight_layout()
            #plt.show()
            pdf.savefig(fig)
            plt.pause(3)
            plt.close()

            print(f'---Figures are saved as {pdfName} in {outPath}')
        # add abfFFolderPath here for data management
        # df_summary = pd.DataFrame(columns=['filename', 'Cm', 'peakSweep', 'peakDensity', 'Rp', 'folder', 'protocol', 'date'])
        df_summary = pd.DataFrame({'filename': abf.abfID,
                                   'Cm': mem_cap,
                                   'peakSweep': peakIndex+1,
                                   'peakDensity': peakValue,
                                   'Rp': Efit,
                                   'folder': os.path.basename(abf.abfFolderPath),
                                   'protocol': abf.protocol,
                                   'date': abf.abfDateTime.date().strftime('%Y%m%d')
                                   })
        # df_summary = df_summary.append(sum_infor, ignore_index=True)


        excelName = os.path.join(outPath, f'{abf.protocol}_{target_vol}.xlsx')
        # load excel, check if it already has this file, if not, write both excel and pdf

        writer = pd.ExcelWriter(excelName, engine='openpyxl')
        # startCol=[0,0,0] # need to create length based on sheetnumber
        if os.path.isfile(excelName): # append data
            writer.book = load_workbook(excelName)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            # check correct file
            # if [sheet.title for sheet in writer.book.worksheets] == ['density','IV','summary']
            # if not been written before
            #density_sheet = f'density_{target_vol}'  # for dif vol
            density_header = [c.value for c in next(writer.book['density'].iter_rows(min_row=1, max_row=1))]

            if abf.abfID not in density_header:
                for df, sheetname in zip([df_density, df_iv], ['density', 'IV']):
                    col_write = [col for col in df.columns if abf.abfID in col]
                    df[col_write].to_excel(writer, sheet_name=sheetname,
                                           startcol=writer.book[sheetname].max_column,
                                           index=False)
                df_summary.to_excel(writer, sheet_name='summary',
                                    startrow=writer.book['summary'].max_row,
                                    header=False,
                                    index=False)
                print(f'---data {abf.abfID} at {target_vol}mV is appended to excel')
            else:
                print(f'---data {abf.abfID} at {target_vol}mV has been added before')

        else: #first time
            for df, sheetname in zip([df_density, df_iv, df_summary], ['density','IV','summary']):
                df.to_excel(writer, sheet_name=sheetname, index=False)
            print(f'    data {abf.abfID} at {target_vol} is written in new excel file {abf.protocol}.xlsx')
        writer.save()

        # save fig as pdf
        # with PdfPages(pdfName) as pdf:
        #     pdf.savefig(fig0)
        #     pdf.savefig(fig)
        #     plt.pause(10)
        #     plt.close('all')
        #     print(f'---Figures are saved as {abf.abfID}.pdf in {outPath}')
        # plt.close('all')
if __name__ == '__main__':
    plt.close('all')
    initFolder = '/Users/yez4/OneDrive - National Institutes of Health/Ephys/Ye_Data/CFTR'
    # save to pdf/excel save with filename as abf.abfFolderPath by sorting files inside genotype folder?
    outPath = '/Users/yez4/OneDrive - National Institutes of Health/Ephys/Output/CFTR'
    if not os.path.isdir(outPath):
        os.mkdir(outPath)

    root = tk.Tk()
    root.withdraw()
    root.update()
    file_paths = filedialog.askopenfilenames(initialdir=initFolder,
                                           title="Select abf file",
                                           filetypes=(("abf files", "*.abf"), ("all files", "*.*")))
    root.update()
    sweepRange = [0, 270] #orai 89, NBC1 89
    baselineSweeps = [] #orai 80-89, NBC 20-29
    # target_vol = -100 # mV, Orai -100, NBC 100
    #make multiple voltages
    target_vols = [-100, 100]

    currentFitTill = 69 # orai fit from peakIndex+3 to 69
    nFit = 4
    for i, file_path in enumerate(file_paths):
        print(f'***Analyzing file {i} of {len(file_paths)-1}:{os.path.basename(file_path)}***')
        debug=1
        if debug:
            runOneAbf(file_path)
        else:
            try:
                runOneAbf(file_path)
            except:
                print(f'---- file {os.path.basename(file_path)} has wrong structure, continue to next')
                continue